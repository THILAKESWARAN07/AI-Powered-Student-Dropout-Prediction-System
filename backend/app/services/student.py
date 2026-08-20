import csv
import io
import json
from typing import List, Dict, Any, Tuple, Optional
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.models.student import (
    Student, StudentAcademics, StudentAttendance, 
    StudentBehaviour, StudentFamily, StudentHealth, 
    StudentTechnology, StudentPrediction
)
from app.models.school import School
from app.schemas.student import ImportSummaryReport, ImportErrorDetail


def parse_csv_or_excel_headers(file_content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]], int]:
    """
    Parses a CSV or Excel file, extracting column headers, preview rows, and total row count.
    """
    headers = []
    rows = []
    
    if filename.endswith(('.xlsx', '.xls')):
        # Load Excel
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = wb.active
        
        # Extract headers
        header_row = next(sheet.iter_rows(values_only=True), None)
        if header_row:
            headers = [str(h).strip() for h in header_row if h is not None]
            
        # Extract preview rows (up to 5) and count total
        total_rows = 0
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                continue
            total_rows += 1
            if len(rows) < 5:
                row_data = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = val
                rows.append({"row_index": total_rows, "data": row_data})
                
    else:
        # Load CSV
        text_content = file_content.decode('utf-8', errors='ignore')
        reader = csv.reader(io.StringIO(text_content))
        
        header_row = next(reader, None)
        if header_row:
            headers = [str(h).strip() for h in header_row if h]
            
        total_rows = 0
        for row in reader:
            if not row:
                continue
            total_rows += 1
            if len(rows) < 5:
                row_data = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = val.strip()
                rows.append({"row_index": total_rows, "data": row_data})
                
    return headers, rows, total_rows


def validate_and_convert_row(
    row_data: Dict[str, Any], 
    mapping: Dict[str, str], 
    row_idx: int,
    db: Session,
    imported_student_ids: set
) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    """
    Validates a single row mapping dataset values to schema elements.
    """
    errors = []
    
    # 1. Reverse-map the dataset keys to model attributes
    model_fields = {}
    for dataset_col, db_col in mapping.items():
        if not dataset_col or not db_col:
            continue
        val = row_data.get(dataset_col)
        model_fields[db_col] = str(val).strip() if val is not None else None

    # Helper function to extract float/int safely
    def get_float(field: str, label: str, required: bool = True, min_val: float = 0.0, max_val: float = 100.0) -> Optional[float]:
        val_str = model_fields.get(field)
        if val_str is None or val_str == "":
            if required:
                errors.append(f"{label} is required.")
            return None
        try:
            val = float(val_str)
            if val < min_val or val > max_val:
                errors.append(f"{label} ({val}) must be between {min_val} and {max_val}.")
                return None
            return val
        except ValueError:
            errors.append(f"{label} ('{val_str}') must be a valid decimal number.")
            return None

    def get_int(field: str, label: str, required: bool = True, min_val: int = 0) -> Optional[int]:
        val_str = model_fields.get(field)
        if val_str is None or val_str == "":
            if required:
                errors.append(f"{label} is required.")
            return None
        try:
            val = int(val_str)
            if val < min_val:
                errors.append(f"{label} ({val}) must be greater than or equal to {min_val}.")
                return None
            return val
        except ValueError:
            errors.append(f"{label} ('{val_str}') must be a valid integer.")
            return None

    def normalize_yes_no(val: Optional[str]) -> Optional[str]:
        if val is None:
            return None
        val_clean = str(val).strip()
        if val_clean.lower() == "yes":
            return "Yes"
        elif val_clean.lower() == "no":
            return "No"
        return val_clean

    # Validate Core Student info
    student_id = model_fields.get("student_id")
    if not student_id:
        errors.append("Student_Id is required.")
    else:
        # Check duplicate within this CSV
        if student_id in imported_student_ids:
            errors.append(f"Duplicate Student_ID '{student_id}' in the import file.")
        else:
            # Check duplicate in database
            exists = db.query(Student).filter(Student.student_id == student_id).first()
            if exists:
                errors.append(f"Student_ID '{student_id}' already exists in database.")

    # Class
    class_name = model_fields.get("class_name")
    if not class_name:
        errors.append("Class is required.")
        
    # Distance to School (km)
    distance = get_float("distance_to_school_km", "Distance to School (km)", required=False, min_val=0.0, max_val=100.0)

    # Transport Mode
    transport = model_fields.get("transport_mode") or "Walking"

    # Travel Time (mins)
    travel_time = get_float("travel_time_min", "Travel Time (mins)", required=False, min_val=0.0, max_val=300.0)

    # Gender
    gender_raw = model_fields.get("gender")
    gender = None
    if not gender_raw:
        errors.append("Gender is required.")
    else:
        gen_clean = gender_raw.strip().lower()
        if gen_clean in ["male", "m"]:
            gender = "Male"
        elif gen_clean in ["female", "f"]:
            gender = "Female"
        else:
            errors.append(f"Gender ('{gender_raw}') must be 'Male' or 'Female'.")
        
    # Age
    age = get_int("age", "Age", required=True, min_val=4)

    # Previous Year Percentage
    prev_pct = get_float("previous_year_percentage", "Previous_Year_Percentage", required=True, min_val=0.0, max_val=100.0)

    # Current Year Percentage
    current_year_pct = get_float("current_year_percentage", "Current_Year_Percentage", required=True, min_val=0.0, max_val=100.0)

    # Overall Percentage
    overall_pct = get_float("overall_percentage", "Overall_Percentage", required=True, min_val=0.0, max_val=100.0)

    # Number of Failures
    failures = get_int("number_of_failures", "Number_of_Failures", required=True, min_val=0)

    # Number of Absences
    absences = get_int("number_of_absences", "Number_of_Absences", required=True, min_val=0)

    # Attendance Percentage
    att_pct = get_float("attendance_percentage", "Attendance_Percentage", required=True, min_val=0.0, max_val=100.0)

    # Attendance Classification
    att_class = model_fields.get("attendance_classification")
    if att_class and att_class not in ["Excellent", "Good", "Moderate", "Poor"]:
        errors.append(f"Attendance_Classification ('{att_class}') must be 'Excellent', 'Good', 'Moderate', or 'Poor'.")

    # Mother and Father Education
    mother_edu = model_fields.get("mother_education") or "Primary"
    father_edu = model_fields.get("father_education") or "Primary"
    parents_edu = mother_edu

    # Family Support
    fam_support = normalize_yes_no(model_fields.get("family_support") or "No")

    # School Support
    sch_support = normalize_yes_no(model_fields.get("school_support") or "No")

    # Internet Access
    internet = normalize_yes_no(model_fields.get("internet_access") or "No")
    if internet not in ["Yes", "No"]:
        errors.append(f"Internet_Access ('{internet}') must be 'Yes' or 'No'.")

    # Health Status
    health_status = model_fields.get("health_status") or "Average"
    if health_status not in ["Good", "Average", "Poor", "Excellent", "Very Poor"]:
        errors.append(f"Health_Status ('{health_status}') must be 'Good', 'Average', 'Poor', 'Excellent', or 'Very Poor'.")

    # Family Income
    income_raw = model_fields.get("family_income")
    income = None
    if not income_raw:
        errors.append("Family_Income is required.")
    else:
        inc_clean = income_raw.strip().lower()
        if inc_clean == "low":
            income = 25000.0
        elif inc_clean == "medium":
            income = 60000.0
        elif inc_clean == "high":
            income = 150000.0
        else:
            try:
                income = float(income_raw)
            except ValueError:
                errors.append(f"Family_Income ('{income_raw}') must be 'Low', 'Medium', 'High', or a decimal number.")

    # Financial Difficulty
    difficulty = normalize_yes_no(model_fields.get("financial_difficulty") or "No")
    if difficulty not in ["Yes", "No"]:
        errors.append(f"Financial_Difficulty ('{difficulty}') must be 'Yes' or 'No'.")

    # Homework Completion
    hw_raw = model_fields.get("homework_completion")
    homework = None
    if not hw_raw:
        errors.append("Homework_Completion is required.")
    else:
        hw_clean = hw_raw.strip().lower()
        if hw_clean == "poor":
            homework = 40.0
        elif hw_clean == "average":
            homework = 65.0
        elif hw_clean == "good":
            homework = 80.0
        elif hw_clean == "excellent":
            homework = 95.0
        else:
            try:
                homework = float(hw_raw)
            except ValueError:
                errors.append(f"Homework_Completion ('{hw_raw}') must be 'Poor', 'Average', 'Good', 'Excellent', or a decimal number.")

    # Low Motivation
    motivation = normalize_yes_no(model_fields.get("low_motivation") or "No")
    if motivation not in ["Yes", "No"]:
        errors.append(f"Low_Motivation ('{motivation}') must be 'Yes' or 'No'.")

    # Mental Health Risk
    mental_raw = model_fields.get("mental_health_risk") or "Low"
    mental_norm = normalize_yes_no(mental_raw)
    if mental_norm in ["Yes", "No"]:
        mental = mental_norm
    else:
        # Check low/medium/high case-insensitively
        mental_clean = mental_raw.strip().lower()
        if mental_clean == "low":
            mental = "Low"
        elif mental_clean == "medium":
            mental = "Medium"
        elif mental_clean == "high":
            mental = "High"
        else:
            mental = mental_raw
            
    if mental not in ["Low", "Medium", "High", "Yes", "No"]:
        errors.append(f"Mental_Health_Risk ('{mental}') must be 'Low', 'Medium', 'High', 'Yes', or 'No'.")

    # Child Labour Risk
    labour = normalize_yes_no(model_fields.get("child_labour_risk") or "No")
    if labour not in ["Yes", "No"]:
        errors.append(f"Child_Labour_Risk ('{labour}') must be 'Yes' or 'No'.")

    # Computer Access
    comp = normalize_yes_no(model_fields.get("computer_access") or "No")
    if comp not in ["Yes", "No"]:
        errors.append(f"Computer_Access ('{comp}') must be 'Yes' or 'No'.")

    # Smartphone Access
    phone = normalize_yes_no(model_fields.get("smartphone_access") or "No")
    if phone not in ["Yes", "No"]:
        errors.append(f"Smartphone_Access ('{phone}') must be 'Yes' or 'No'.")

    # Electricity Availability
    elec = normalize_yes_no(model_fields.get("electricity_availability") or "Yes")
    if elec not in ["Yes", "No"]:
        errors.append(f"Electricity_Availability ('{elec}') must be 'Yes' or 'No'.")

    if errors:
        return None, errors

    # Package objects
    mapped_data = {
        "student": Student(
            student_id=student_id,
            full_name=f"Student {student_id}",
            gender=gender,
            age=age,
            class_name=class_name,
            section="A",
            medium_of_instruction="English",
            community="General",
            distance_to_school_km=distance,
            transport_mode=transport,
            travel_time_min=travel_time,
            school_type="Government",
            teacher_student_ratio="1:35"
        ),
        "academics": StudentAcademics(
            previous_year_percentage=prev_pct,
            unit_test_average=current_year_pct,
            quarterly_exam=current_year_pct,
            half_yearly_exam=current_year_pct,
            annual_exam=current_year_pct,
            mathematics_marks=current_year_pct,
            science_marks=current_year_pct,
            english_marks=current_year_pct,
            social_science_marks=current_year_pct,
            regional_language_marks=current_year_pct,
            overall_percentage=current_year_pct,
            number_of_failed_subjects=failures,
            academic_backlogs="Yes" if failures > 0 else "No"
        ),
        "attendance": StudentAttendance(
            attendance_percentage=att_pct,
            consecutive_absences=absences,
            leave_days=0,
            late_arrivals=0
        ),
        "behaviour": StudentBehaviour(
            homework_completion=homework,
            assignment_submission_rate=homework,
            classroom_participation="Medium",
            discipline_incidents=0,
            teacher_feedback=sch_support,
            participation_in_extracurricular="No",
            library_usage="Medium",
            low_motivation=motivation,
            bullying_experience="No"
        ),
        "family": StudentFamily(
            family_income=income,
            parents_education=parents_edu,
            parents_occupation="Farmer",
            single_parent="No",
            number_of_siblings=0,
            guardian_support=fam_support,
            home_study_hours=3.0,
            financial_difficulty=difficulty,
            child_labour_risk=labour,
            frequent_migration="No",
            family_issues="No"
        ),
        "health": StudentHealth(
            chronic_illness="No",
            nutrition_status=health_status,
            vision_problems="No",
            mental_health_risk=mental,
            disability_status="No",
            midday_meal_beneficiary="No"
        ),
        "technology": StudentTechnology(
            internet_access=internet,
            smartphone_access=phone,
            computer_access=comp,
            electricity_availability=elec
        ),
        "prediction": StudentPrediction(
            dropout_risk="Low",
            dropout_status="No"
        )
    }
    
    return mapped_data, []


def import_mapped_records(
    db: Session,
    file_content: bytes,
    filename: str,
    mapping: Dict[str, str],
    school_id: int
) -> ImportSummaryReport:
    """
    Iterates over the import sheet, validates row schemas, maps fields, and writes to database.
    """
    total_records = 0
    imported = 0
    skipped = 0
    failed = 0
    duplicates = 0
    error_list = []
    
    imported_student_ids = set()
    failed_rows = []

    # Verify target school exists
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise ValueError(f"School with ID {school_id} does not exist.")

    # 1. Parse rows
    headers = []
    rows_raw = []
    
    if filename.endswith(('.xlsx', '.xls')):
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        sheet = wb.active
        header_row = next(sheet.iter_rows(values_only=True), None)
        if header_row:
            headers = [str(h).strip() for h in header_row if h is not None]
        for row in sheet.iter_rows(values_only=True):
            if not rows_raw and len(row) > 0: # skip first row (header)
                rows_raw.append(row) # dummy push to skip header
                continue
            row_data = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = val
            rows_raw.append(row_data)
        rows_raw = rows_raw[1:] # strip header
    else:
        text_content = file_content.decode('utf-8', errors='ignore')
        reader = csv.reader(io.StringIO(text_content))
        header_row = next(reader, None)
        if header_row:
            headers = [str(h).strip() for h in header_row if h]
        for row in reader:
            if not row:
                continue
            row_data = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = val.strip()
            rows_raw.append(row_data)

    total_records = len(rows_raw)

    # 1.1 Automatic column detection if mapping is empty
    expected_db_cols = [
        'student_id', 'class_name', 'distance_to_school_km', 'transport_mode', 'travel_time_min', 'gender', 'age',
        'previous_year_percentage', 'current_year_percentage', 'overall_percentage', 'number_of_failures', 'number_of_absences',
        'attendance_percentage', 'attendance_classification', 'mother_education', 'father_education',
        'family_support', 'school_support', 'internet_access', 'health_status', 'family_income', 'financial_difficulty',
        'homework_completion', 'low_motivation', 'mental_health_risk', 'child_labour_risk', 'computer_access', 'smartphone_access',
        'electricity_availability'
    ]

    def normalize_name(name: str) -> str:
        import re
        return re.sub(r'[^a-z0-9]', '', name.lower())

    if not mapping:
        mapping = {}

    mapped_values = set(mapping.values())
    for header in headers:
        if header in mapping:
            continue
        norm_header = normalize_name(header)
        for col in expected_db_cols:
            if col in mapped_values:
                continue
            norm_col = normalize_name(col)
            if norm_col == "classname" and norm_header == "class":
                mapping[header] = col
                mapped_values.add(col)
                break
            elif norm_col == "traveltimemin" and norm_header == "traveltimemins":
                mapping[header] = col
                mapped_values.add(col)
                break
            elif norm_col == norm_header:
                mapping[header] = col
                mapped_values.add(col)
                break

    # 1.2 Column Validation (reject invalid datasets gracefully)
    REQUIRED_FIELDS = {
        "student_id": "Student_Id",
        "class_name": "Class",
        "gender": "Gender",
        "age": "Age",
        "previous_year_percentage": "Previous_Year_Percentage",
        "current_year_percentage": "Current_Year_Percentage",
        "overall_percentage": "Overall_Percentage",
        "number_of_failures": "Number_of_Failures",
        "number_of_absences": "Number_of_Absences",
        "attendance_percentage": "Attendance_Percentage",
        "homework_completion": "Homework_Completion",
        "family_income": "Family_Income"
    }

    mapped_db_cols = set(mapping.values())
    missing_required = [label for field, label in REQUIRED_FIELDS.items() if field not in mapped_db_cols]
    if missing_required:
        raise ValueError(f"Missing required columns in mapping: {', '.join(missing_required)}")

    # 2. Iterate and validate
    for idx, row_data in enumerate(rows_raw):
        row_idx = idx + 1
        
        # Execute validation
        mapped_data, row_errors = validate_and_convert_row(
            row_data=row_data,
            mapping=mapping,
            row_idx=row_idx,
            db=db,
            imported_student_ids=imported_student_ids
        )
        
        if row_errors:
            failed += 1
            error_list.append(ImportErrorDetail(
                row_index=row_idx,
                student_id=row_data.get("Student_ID") or row_data.get("Student_Id") or row_data.get("student_id"),
                errors=row_errors
            ))
            
            # Record failed row data for download
            failed_row_record = dict(row_data)
            failed_row_record["_row_index"] = row_idx
            failed_row_record["_errors"] = ", ".join(row_errors)
            failed_rows.append(failed_row_record)
            
            # Count duplicates if it was a duplicate error
            if any("already exists" in e or "Duplicate Student_ID" in e for e in row_errors):
                duplicates += 1
            continue

        # Save to database
        try:
            student = mapped_data["student"]
            student.school_id = school_id
            
            # Save core student record
            db.add(student)
            db.flush() # populates student.id
            
            # Assign relations
            academics = mapped_data["academics"]
            academics.student_id = student.id
            db.add(academics)
            
            attendance = mapped_data["attendance"]
            attendance.student_id = student.id
            db.add(attendance)
            
            behaviour = mapped_data["behaviour"]
            behaviour.student_id = student.id
            db.add(behaviour)
            
            family = mapped_data["family"]
            family.student_id = student.id
            db.add(family)
            
            health = mapped_data["health"]
            health.student_id = student.id
            db.add(health)
            
            tech = mapped_data["technology"]
            tech.student_id = student.id
            db.add(tech)
            
            pred = mapped_data["prediction"]
            pred.student_id = student.id
            db.add(pred)
            
            db.commit()

            # Execute model auto-prediction immediately
            try:
                from app.services.prediction_service import predict_student
                predict_student(db, student.id)
            except Exception as pred_err:
                from app.core.logging import logger
                logger.error(f"Failed to auto-predict for imported student {student.id}: {str(pred_err)}")
            
            imported += 1
            imported_student_ids.add(student.student_id)
            
        except Exception as e:
            db.rollback()
            failed += 1
            error_msg = f"Database write error: {str(e)}"
            error_list.append(ImportErrorDetail(
                row_index=row_idx,
                student_id=row_data.get("Student_ID") or row_data.get("Student_Id") or row_data.get("student_id"),
                errors=[error_msg]
            ))
            failed_row_record = dict(row_data)
            failed_row_record["_row_index"] = row_idx
            failed_row_record["_errors"] = error_msg
            failed_rows.append(failed_row_record)

    failed_json = json.dumps(failed_rows) if failed_rows else None

    return ImportSummaryReport(
        total_records=total_records,
        imported=imported,
        skipped=skipped,
        failed=failed,
        duplicates=duplicates,
        errors=error_list,
        failed_rows_json=failed_json
    )


